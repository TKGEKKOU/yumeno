from __future__ import annotations

import csv
import hashlib
import json
import sqlite3
from pathlib import Path
from typing import Any

from structured_data.contracts import (
    StructuredColumn,
    StructuredImportResult,
    StructuredTable,
)


MAX_SHEETS = 20
MAX_ROWS = 100_000
MAX_COLUMNS = 100
MAX_CELL_CHARS = 10_000


def _physical_table(document_id: str, sheet_index: int) -> str:
    digest = hashlib.sha256(f"{document_id}:{sheet_index}".encode()).hexdigest()[:16]
    return f"t_{digest}"


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    text = str(value)
    return text[:MAX_CELL_CHARS]


def _read_csv(path: Path) -> list[tuple[str, list[str], list[list[Any]]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = []
        for index, row in enumerate(reader):
            if index > MAX_ROWS:
                raise ValueError("structured_row_limit_exceeded")
            if len(row) > MAX_COLUMNS:
                raise ValueError("structured_column_limit_exceeded")
            rows.append([_normalize_cell(value) for value in row])
    if not rows:
        raise ValueError("structured_file_is_empty")
    headers = [str(value or f"column_{index + 1}") for index, value in enumerate(rows[0])]
    width = len(headers)
    data = [(row + [None] * width)[:width] for row in rows[1:]]
    return [(path.stem, headers, data)]


def _read_xlsx(path: Path) -> list[tuple[str, list[str], list[list[Any]]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    output = []
    try:
        if len(workbook.sheetnames) > MAX_SHEETS:
            raise ValueError("structured_sheet_limit_exceeded")
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            iterator = sheet.iter_rows(values_only=True)
            first = next(iterator, None)
            if first is None:
                continue
            if len(first) > MAX_COLUMNS:
                raise ValueError("structured_column_limit_exceeded")
            headers = [
                str(value or f"column_{index + 1}")
                for index, value in enumerate(first)
            ]
            data = []
            for index, row in enumerate(iterator, start=1):
                if index > MAX_ROWS:
                    raise ValueError("structured_row_limit_exceeded")
                data.append([_normalize_cell(value) for value in row[: len(headers)]])
            output.append((sheet_name, headers, data))
    finally:
        workbook.close()
    if not output:
        raise ValueError("structured_file_is_empty")
    return output


def _column_type(values: list[Any]) -> str:
    present = [value for value in values if value not in (None, "")]
    if not present:
        return "TEXT"
    try:
        if all(str(value).strip().lstrip("+-").isdigit() for value in present):
            return "INTEGER"
        if all(_is_float(value) for value in present):
            return "REAL"
    except (TypeError, ValueError):
        pass
    return "TEXT"


def _is_float(value: Any) -> bool:
    float(value)
    return True


def _coerce(value: Any, data_type: str) -> Any:
    if value in (None, ""):
        return None
    if data_type == "INTEGER":
        return int(value)
    if data_type == "REAL":
        return float(value)
    return str(value)


def _schema_card(tables: list[StructuredTable]) -> str:
    lines = ["# Structured data schema"]
    for table in tables:
        lines.append(f"\n## {table.display_name} ({table.physical_name})")
        lines.append(f"Rows: {table.row_count}")
        for column in table.columns:
            lines.append(
                f"- {column.display_name} -> {column.physical_name} [{column.data_type}]"
            )
    return "\n".join(lines)


def import_structured_file(
    source_path: Path,
    *,
    db_path: Path,
    workspace_id: str,
    knowledge_space_id: str,
    document_id: str,
) -> StructuredImportResult:
    suffix = source_path.suffix.lower()
    if suffix == ".csv":
        sheets = _read_csv(source_path)
    elif suffix == ".xlsx":
        sheets = _read_xlsx(source_path)
    else:
        raise ValueError("unsupported_structured_format")

    db_path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(db_path)
    tables: list[StructuredTable] = []
    try:
        connection.execute(
            "CREATE TABLE IF NOT EXISTS _yumeno_datasets ("
            "physical_name TEXT PRIMARY KEY, document_id TEXT NOT NULL, "
            "workspace_id TEXT NOT NULL, knowledge_space_id TEXT NOT NULL, "
            "display_name TEXT NOT NULL, schema_json TEXT NOT NULL, row_count INTEGER NOT NULL)"
        )
        previous = connection.execute(
            "SELECT physical_name FROM _yumeno_datasets WHERE document_id = ?",
            (document_id,),
        ).fetchall()
        for (table_name,) in previous:
            if table_name.startswith("t_"):
                connection.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        connection.execute(
            "DELETE FROM _yumeno_datasets WHERE document_id = ?", (document_id,)
        )

        for sheet_index, (sheet_name, headers, rows) in enumerate(sheets):
            physical_name = _physical_table(document_id, sheet_index)
            width = len(headers)
            column_types = [
                _column_type([row[index] if index < len(row) else None for row in rows[:1000]])
                for index in range(width)
            ]
            columns = tuple(
                StructuredColumn(f"c_{index + 1:03d}", headers[index], column_types[index])
                for index in range(width)
            )
            definition = ", ".join(
                f'"{column.physical_name}" {column.data_type}' for column in columns
            )
            connection.execute(f'CREATE TABLE "{physical_name}" ({definition})')
            placeholders = ", ".join("?" for _ in columns)
            if rows:
                connection.executemany(
                    f'INSERT INTO "{physical_name}" VALUES ({placeholders})',
                    [
                        tuple(
                            _coerce(row[index] if index < len(row) else None, column_types[index])
                            for index in range(width)
                        )
                        for row in rows
                    ],
                )
            table = StructuredTable(physical_name, sheet_name, columns, len(rows))
            tables.append(table)
            connection.execute(
                "INSERT INTO _yumeno_datasets VALUES (?, ?, ?, ?, ?, ?, ?)",
                (
                    physical_name,
                    document_id,
                    workspace_id,
                    knowledge_space_id,
                    sheet_name,
                    json.dumps([column.__dict__ for column in columns], ensure_ascii=False),
                    len(rows),
                ),
            )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()
    return StructuredImportResult(
        document_id=document_id,
        tables=tuple(tables),
        row_count=sum(table.row_count for table in tables),
        schema_card=_schema_card(tables),
    )
